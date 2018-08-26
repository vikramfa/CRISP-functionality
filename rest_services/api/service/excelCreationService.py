import json
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Side,PatternFill,Font, colors
import datetime
import pytz
from django.conf import settings
from collections import OrderedDict

class ExcelCreationService:

    def getXlsxFieldsAlignmentByTopic(self, topic_name):
        cols_array_align = None
        if(topic_name == 'domainiq'):
            cols_array_align = {'A':4,'B':10,'C':22,'D':10,'E':15,'F':15,'G':8,'H':15,'I':20,'J':25,'K':10,
                                'L': 12,'M':6,'N':12,'O':15,'P':18,'Q':10,'R':20,'S':30,'T':30,'U':13}
        elif(topic_name == 'shodan'):
            cols_array_align = {'A':4,'B':30,'C':22,'D':12,'E':14,'F':14,'G':25,'H':15,'I':15,'J':30,
                                'K': 14,'L':14,'M':25,'N':12,'O':12,'P':15,'Q':15,'R':10,'S':30,'T':15,'U':15}
        elif (topic_name == 'census'):
            cols_array_align = {'A': 4, 'B': 25, 'C':22,'D': 15, 'E': 12, 'F': 14, 'G': 14, 'H': 10, 'I': 20, 'J': 20, 'K': 25,
                                'L': 13, 'M': 13, 'N': 10, 'O': 18, 'P': 18, 'Q': 18, 'R': 12, 'S': 15, 'T': 30,
                                'U': 30, 'V': 45}
        elif (topic_name == 'censys'):
            cols_array_align = {'A': 4, 'B': 25, 'C':22,'D': 24, 'E': 20, 'F': 20, 'G': 20, 'H': 20, 'I': 20, 'J': 20, 'K': 25,
                                'L': 20, 'M': 20, 'N': 20, 'O': 20, 'P': 20, 'Q': 20, 'R': 20, 'S': 20, 'T': 20,
                                'U': 20, 'V': 20, 'W':20, 'X':20, 'Y':20, 'Z':20, 'AA':20, 'AB': 20, 'AC': 20, 'AD': 20, 'AE': 20,
                                'AF': 20, 'AG': 20, 'AH': 20, 'AI': 20, 'AJ': 20, 'AK': 20,
                                'AL': 20, 'AM': 20, 'AN': 20, 'AO': 20, 'AP': 20, 'AQ': 20, 'AR': 20, 'AS': 20, 'AT': 20,
                                'AU': 20, 'AV': 20}

        return  cols_array_align

    def setDataAvail(self, dictData):
        dataList = []
        dataAvailList = []
        for index,element in enumerate(dictData):
            for indx, col in enumerate(element.keys()):
                if indx > 1:
                    dataPresent = False if element[col] == '' or element[col] is None else True

            datPresentVal = 'Yes' if dataPresent else 'No'
            dataAvailList.append(datPresentVal)

        for index, element in enumerate(dictData):
            dictDataCpy = OrderedDict()
            for indx, (key, val) in enumerate(element.items()):
                if(indx==2):
                    dictDataCpy['Data Availability'] = dataAvailList[index]
                    dictDataCpy[key] = val
                else:
                    dictDataCpy[key] = val
            dataList.append(dictDataCpy)

        return dataList


    def createFinalXlsFile(self, result_file_name, batch_obj):
        headers = ['Batch Name:', 'File Upload Time:', 'Number of Failed Records - Shodan:',
                   'Number of Failed Records - Censys:', 'Number of Failed Records - DomainIQ:',
                   'Number of Record Processed:',
                   'Number of Records with no data available - Shodan',
                   'Number of Records with no data available - Cenys',
                   'Number of Records with no data available - DomainIQ', 'Record Processing Time:',
                   'Uploader Name:']
        rightAlignment = Alignment(horizontal='right',
                                   vertical='bottom',
                                   text_rotation=0,
                                   wrap_text=True,
                                   shrink_to_fit=False,
                                   indent=0)
        leftAlignment = Alignment(horizontal='left',
                                  vertical='bottom',
                                  text_rotation=0,
                                  wrap_text=True,
                                  shrink_to_fit=False,
                                  indent=0)
        centerAlignment = Alignment(horizontal='center',
                                    vertical='bottom',
                                    text_rotation=0,
                                    wrap_text=True,
                                    shrink_to_fit=False,
                                    indent=0)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", bgColor="FFFFFF")
        ft = Font(color=colors.WHITE)
        wb = Workbook()
        ws = wb.active
        # colAlignment = ColumnDimension(ws, bestFit=True)
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 35
        ws.title = 'Batch Details'
        ws.merge_cells('A1:B1')
        headerCell = ws.cell(row=1, column=1, value="BATCH DETAILS")
        headerCell.fill = fill
        headerCell.font = Font(color=colors.WHITE, bold=True)
        headerCell.alignment = centerAlignment
        row_index = 2

        now = datetime.datetime.now()
        time_diff = str(pytz.UTC.localize(now) - batch_obj.created_at).split(':')
        time_diff_format = '%s Hours and %s Minutes' % (time_diff[0], time_diff[1])
        data = [batch_obj.batch_name,
                batch_obj.created_at.strftime("%b %d %Y at %I:%M %p"),
                str(batch_obj.shodan_failed_records),
                str(batch_obj.censys_failed_records),
                str(batch_obj.domainiq_failed_records),
                str(batch_obj.shodan_processed_records + batch_obj.domainiq_processed_records +
                    batch_obj.censys_processed_records),
                "0", "0", "0",
                time_diff_format,
                str(batch_obj.user_created_by)
                ]
        for i in range(0, len(headers)):
            column_index = 1
            cell = ws.cell(row=row_index, column=column_index, value=headers[i])
            cell.alignment = rightAlignment
            cell.border = border
            cell.fill = fill
            cell.font = ft
            cell = ws.cell(row=row_index, column=(column_index + 1), value=data[i])
            cell.alignment = leftAlignment
            cell.border = border
            row_index += 1
        # wb.save(result_file_name)

        for topic in ('shodan', 'censys', 'domainiq'):
            file_name = '%s/result_files/batch_%s_%s.json' % (settings.MEDIA_ROOT, batch_obj.batch_id, topic)
            json_file = open(file_name, 'r')
            data1 = json.load(json_file, object_pairs_hook=OrderedDict)
            ws = wb.create_sheet(topic)
            try:
                for row in data1:
                    del row['batch_id']
                data = self.setDataAvail(data1)
                row_index = 1
                for indx, col in enumerate(data[0].keys()):
                        headerCell = ws.cell(row=row_index, column=(1 + indx), value=col)
                        headerCell.fill = fill
                        headerCell.font = Font(color=colors.WHITE, bold=True)
                        headerCell.alignment = leftAlignment
                row_index = 2
                for element in data:
                    for indx,col in enumerate(element.keys()):
                        cell = ws.cell(row=row_index, column=(1 + indx), value=element[col])
                        cell.alignment = leftAlignment
                        cell.border = border
                    row_index += 1

                    cols_array_align = self.getXlsxFieldsAlignmentByTopic(topic)
                    if cols_array_align is not None:
                        for key,value in cols_array_align.items():
                            ws.column_dimensions[key].width = value

            except FileNotFoundError:
                print("\"%s\" partial file not found" % topic)
        wb.save(result_file_name)